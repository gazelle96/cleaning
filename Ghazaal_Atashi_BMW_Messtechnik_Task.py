import openpyxl
import pandas as pd
import os

# Prompt the user for the path to the Excel file
excel_file = input("Please enter the path to your Excel file (e.g., C:\\Users\\ASUS\\Downloads\\Sample_list.xlsx): ")

# Load the workbook and the first sheet
wb = openpyxl.load_workbook(excel_file)
ws = wb.active

# Unmerging cells and propagate their values separately
def unmerge_and_propagate(ws):
    merged_ranges = list(ws.merged_cells.ranges)
    for merged_cell in merged_ranges:
        min_col, min_row, max_col, max_row = merged_cell.min_col, merged_cell.min_row, merged_cell.max_col, merged_cell.max_row
        top_left_value = ws.cell(row=min_row, column=min_col).value

        ws.unmerge_cells(start_row=min_row, start_column=min_col, end_row=max_row, end_column=max_col)

        if top_left_value is not None and top_left_value != '':
            for row in range(min_row, max_row + 1):
                for col in range(min_col, max_col + 1):
                    ws.cell(row=row, column=col).value = top_left_value

unmerge_and_propagate(ws)

# Construct the output file path in the same directory as the original file
directory = os.path.dirname(excel_file)
output_file = os.path.join(directory, 'unmerged_file.xlsx')

wb.save(output_file)

df = pd.read_excel(output_file, engine='openpyxl')

# Check and convert "Date of birth" to the desired format (dd/mm/yyyy)
def convert_dob_format(date):
    if isinstance(date, pd.Timestamp):
        return date.strftime('%d/%m/%Y')
    try:
        date_obj = pd.to_datetime(date, errors='coerce')
        if pd.isnull(date_obj):
            return date  # If parsing fails, return the original value
        return date_obj.strftime('%d/%m/%Y')
    except Exception:
        return date

if 'Date of birth' in df.columns:
    df['Date of birth'] = df['Date of birth'].apply(convert_dob_format)

df.to_excel(output_file, index=False, engine='openpyxl')

# Loop until a valid ID is entered
while True:
    search_id = input("Which ID are you looking for? ")

    if search_id in df['ID'].values:
        break
    else:
        print("This ID does not exist! Please try another one: ")

# Prompt the user for their choice of cleaning the data or not
print("Please choose one of the options below:")
print("1: Look for that ID without cleaning the data")
print("2: Look for that ID after cleaning the data")
choice = input("Enter 1 or 2: ")

if choice == '2':
    cleaned_df = df.dropna()
    filtered_df = cleaned_df[cleaned_df['ID'] == search_id]
else:
    filtered_df = df[df['ID'] == search_id]

# Construct the output file path for the filtered data and save filtered data to a new Excel file
filtered_output_file = os.path.join(directory, f'filtered_{search_id}.xlsx')
filtered_df.to_excel(filtered_output_file, index=False, engine='openpyxl')

print(f"Unmerged and updated file saved at: {output_file}")
print(f"Filtered file for ID {search_id} saved at: {filtered_output_file}")


